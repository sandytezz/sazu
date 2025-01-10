import  time
import openpyxl
import warnings
from datetime import datetime
warnings.simplefilter(action='ignore', category=UserWarning)
wb = openpyxl.load_workbook("D:\\PYTHON\\fileconversion\\David Horowitz\\Social_Circle_Nursing_and _Rehab\\ACA60346354SOCIALCIRCFINALreformatted.xlsx")
worksheet = wb["ACA60346354SOCIALCIRCFINALrefor"]
Country = "USA"
Part_three = "Y"
Plan_start_month = "01-January"
Self = "Yes"

write_row=3
total_increase=0
for i in range(3,106):
    SSN = worksheet.cell(row=i, column=1).value
    firstname = worksheet.cell(row=i, column=2).value
    Middle_name = worksheet.cell(row=i, column=3).value
    lastname = worksheet.cell(row=i, column=4).value
    sufix = worksheet.cell(row=i, column=5).value
    DOB = worksheet.cell(row=i, column=6).value
    datetime_string = str(DOB)
    date_fr = datetime_string.split(' ')[0]
    print(date_fr)
    
    Address = worksheet.cell(row=i, column=7).value
    city = worksheet.cell(row=i, column=8).value
    state = worksheet.cell(row=i, column=9).value
    zipcode = worksheet.cell(row=i, column=10).value

    # line 14 codes

    line14_all_12_months = worksheet.cell(row=i, column=12).value
    One_JAN = worksheet.cell(row=i, column=13).value
    One_FEB = worksheet.cell(row=i, column=14).value
    One_MAR = worksheet.cell(row=i, column=15).value
    One_APRL = worksheet.cell(row=i, column=16).value
    One_MAY = worksheet.cell(row=i, column=17).value
    One_JUNE = worksheet.cell(row=i, column=18).value
    One_JULY = worksheet.cell(row=i, column=19).value
    One_AUG = worksheet.cell(row=i, column=20).value
    One_SEP = worksheet.cell(row=i, column=21).value
    One_OCT = worksheet.cell(row=i, column=22).value
    One_NOV = worksheet.cell(row=i, column=23).value
    One_DEC = worksheet.cell(row=i, column=24).value

    # premium

    Premium_all_12_months = worksheet.cell(row=i, column=25).value
    PJAN = worksheet.cell(row=i, column=26).value
    PFEB = worksheet.cell(row=i, column=27).value
    PMAR = worksheet.cell(row=i, column=28).value
    PAPRL = worksheet.cell(row=i, column=29).value
    PMAY = worksheet.cell(row=i, column=30).value
    PJUNE = worksheet.cell(row=i, column=31).value
    PJULY = worksheet.cell(row=i, column=32).value
    PAUG = worksheet.cell(row=i, column=33).value
    PSEP = worksheet.cell(row=i, column=34).value
    POCT = worksheet.cell(row=i, column=35).value
    PNOV = worksheet.cell(row=i, column=36).value
    PDEC = worksheet.cell(row=i, column=37).value

    # line 16 codes

    line16_all_12_months = worksheet.cell(row=i, column=38).value
    Two_JAN = worksheet.cell(row=i, column=39).value
    Two_FEB = worksheet.cell(row=i, column=40).value
    Two_MAR = worksheet.cell(row=i, column=41).value
    Two_APRL = worksheet.cell(row=i, column=42).value
    Two_MAY = worksheet.cell(row=i, column=43).value
    Two_JUNE = worksheet.cell(row=i, column=44).value
    Two_JULY = worksheet.cell(row=i, column=45).value
    Two_AUG = worksheet.cell(row=i, column=46).value
    Two_SEP = worksheet.cell(row=i, column=47).value
    Two_OCT = worksheet.cell(row=i, column=48).value
    Two_NOV = worksheet.cell(row=i, column=49).value
    Two_DEC = worksheet.cell(row=i, column=50).value

    dependent_count = 0

    dep_1 = worksheet.cell(row=i, column=65).value
    dep_2 = worksheet.cell(row=i, column=84).value
    dep_3 = worksheet.cell(row=i, column=103).value
    dep_4 = worksheet.cell(row=i, column=122).value
    dep_5 = worksheet.cell(row=i, column=141).value
    dep_6 = worksheet.cell(row=i, column=160).value
    dep_7 = worksheet.cell(row=i, column=179).value
    dep_8 = worksheet.cell(row=i, column=198).value

    # part three

    Covered_Part = worksheet.cell(row=i, column=51).value
    Dep = worksheet.cell(row=i, column=64).value
    

    wb2 = openpyxl.load_workbook("D:\\PYTHON\\fileconversion\\David Horowitz\\Social_Circle_Nursing_and _Rehab\\ACAwise_FullService_Core_1094and1095-C_2024_V1.0.xlsx")
    ws2 = wb2["1095-C Data"]
    ws2.cell(row=write_row, column=5).value = SSN
    ws2.cell(row=write_row, column=1).value = firstname
    ws2.cell(row=write_row, column=3).value = lastname
    ws2.cell(row=write_row, column=4).value = sufix
    ws2.cell(row=write_row, column=2).value = Middle_name
    ws2.cell(row=write_row, column=6).value = Address
    ws2.cell(row=write_row, column=8).value = city
    ws2.cell(row=write_row, column=9).value = state
    ws2.cell(row=write_row, column=10).value = zipcode
    ws2.cell(row=write_row, column=11).value = Country
    ws2.cell(row=write_row, column=17).value = Plan_start_month

    # line 14 codes

    ws2.cell(row=write_row, column=18).value = line14_all_12_months
    ws2.cell(row=write_row, column=19).value = One_JAN
    ws2.cell(row=write_row, column=20).value = One_FEB
    ws2.cell(row=write_row, column=21).value = One_MAR
    ws2.cell(row=write_row, column=22).value = One_APRL
    ws2.cell(row=write_row, column=23).value = One_MAY
    ws2.cell(row=write_row, column=24).value = One_JUNE
    ws2.cell(row=write_row, column=25).value = One_JULY
    ws2.cell(row=write_row, column=26).value = One_AUG
    ws2.cell(row=write_row, column=27).value = One_SEP
    ws2.cell(row=write_row, column=28).value = One_OCT
    ws2.cell(row=write_row, column=29).value = One_NOV
    ws2.cell(row=write_row, column=30).value = One_DEC

    # premium

    ws2.cell(row=write_row, column=31).value = Premium_all_12_months
    ws2.cell(row=write_row, column=32).value = PJAN
    ws2.cell(row=write_row, column=33).value = PFEB
    ws2.cell(row=write_row, column=34).value = PMAR
    ws2.cell(row=write_row, column=35).value = PAPRL
    ws2.cell(row=write_row, column=36).value = PMAY
    ws2.cell(row=write_row, column=37).value = PJUNE
    ws2.cell(row=write_row, column=38).value = PJULY
    ws2.cell(row=write_row, column=39).value = PAUG
    ws2.cell(row=write_row, column=40).value = PSEP
    ws2.cell(row=write_row, column=41).value = POCT
    ws2.cell(row=write_row, column=42).value = PNOV
    ws2.cell(row=write_row, column=43).value = PDEC

    # line 16 codes

    ws2.cell(row=write_row, column=44).value = line16_all_12_months
    ws2.cell(row=write_row, column=45).value = Two_JAN
    ws2.cell(row=write_row, column=46).value = Two_FEB
    ws2.cell(row=write_row, column=47).value = Two_MAR
    ws2.cell(row=write_row, column=48).value = Two_APRL
    ws2.cell(row=write_row, column=49).value = Two_MAY
    ws2.cell(row=write_row, column=50).value = Two_JUNE
    ws2.cell(row=write_row, column=51).value = Two_JULY
    ws2.cell(row=write_row, column=52).value = Two_AUG
    ws2.cell(row=write_row, column=53).value = Two_SEP
    ws2.cell(row=write_row, column=54).value = Two_OCT
    ws2.cell(row=write_row, column=55).value = Two_NOV
    ws2.cell(row=write_row, column=56).value = Two_DEC

    if Covered_Part == "Y":
        if ws2.cell(row=write_row, column=38).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=77).value = "Yes"

        else:
            ws2.cell(row=write_row, column=77).value = "No"

        if ws2.cell(row=write_row, column=45).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=78).value = "Yes"
        else:
            ws2.cell(row=write_row, column=78).value = "No"

        if ws2.cell(row=write_row, column=46).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=79).value = "Yes"
        else:
            ws2.cell(row=write_row, column=79).value = "No"

        if ws2.cell(row=write_row, column=47).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=80).value = "Yes"
        else:
            ws2.cell(row=write_row, column=80).value = "No"

        if ws2.cell(row=write_row, column=48).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=81).value = "Yes"
        else:
            ws2.cell(row=write_row, column=81).value = "No"

        if ws2.cell(row=write_row, column=49).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=82).value = "Yes"
        else:
            ws2.cell(row=write_row, column=82).value = "No"

        if ws2.cell(row=write_row, column=50).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=83).value = "Yes"
        else:
            ws2.cell(row=write_row, column=83).value = "No"

        if ws2.cell(row=write_row, column=51).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=84).value = "Yes"
        else:
            ws2.cell(row=write_row, column=84).value = "No"

        if ws2.cell(row=write_row, column=52).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=85).value = "Yes"
        else:
            ws2.cell(row=write_row, column=85).value = "No"

        if ws2.cell(row=write_row, column=53).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=86).value = "Yes"
        else:
            ws2.cell(row=write_row, column=86).value = "No"

        if ws2.cell(row=write_row, column=54).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=87).value = "Yes"
        else:
            ws2.cell(row=write_row, column=87).value = "No"

        if ws2.cell(row=write_row, column=55).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=88).value = "Yes"
        else:
            ws2.cell(row=write_row, column=88).value = "No"

        if ws2.cell(row=write_row, column=56).value == "2C":
            ws2.cell(row=write_row, column=70).value = Self
            ws2.cell(row=write_row, column=71).value = firstname
            ws2.cell(row=write_row, column=72).value = Middle_name
            ws2.cell(row=write_row, column=73).value = lastname
            ws2.cell(row=write_row, column=74).value = sufix
            ws2.cell(row=write_row, column=75).value = SSN
            ws2.cell(row=write_row, column=76).value = date_fr
            ws2.cell(row=write_row, column=89).value = "Yes"
        else:
            ws2.cell(row=write_row, column=89).value = "No"

    
    dep = [dep_1, dep_2, dep_3, dep_4, dep_5, dep_6, dep_7, dep_8]

    dep_1_firstname = worksheet.cell(row=i, column=65).value
    dep_1_middlename = worksheet.cell(row=i, column=66).value
    dep_1_lastname = worksheet.cell(row=i, column=67).value
    dep_1_ssn = worksheet.cell(row=i, column=68).value
    dep_1_s_dob = worksheet.cell(row=i, column=69).value
    date_string = str(dep_1_s_dob)
    dep_1_dob = datetime_string.split(' ')[0]

    dep_2_firstname = worksheet.cell(row=i, column=84).value
    dep_2_middlename = worksheet.cell(row=i, column=85).value
    dep_2_lastname = worksheet.cell(row=i, column=86).value
    dep_2_ssn = worksheet.cell(row=i, column=87).value
    dep_2_s_dob = worksheet.cell(row=i, column=68).value
    date_string = str(dep_2_s_dob)
    dep_2_dob = datetime_string.split(' ')[0]

    dep_3_firstname = worksheet.cell(row=i, column=103).value
    dep_3_middlename = worksheet.cell(row=i, column=104).value
    dep_3_lastname = worksheet.cell(row=i, column=105).value
    dep_3_ssn = worksheet.cell(row=i, column=106).value
    dep_3_s_dob = worksheet.cell(row=i, column=107).value
    date_string = str(dep_3_s_dob)
    dep_3_dob = datetime_string.split(' ')[0]

    dep_4_firstname = worksheet.cell(row=i, column=122).value
    dep_4_middlename = worksheet.cell(row=i, column=123).value
    dep_4_lastname = worksheet.cell(row=i, column=124).value
    dep_4_ssn = worksheet.cell(row=i, column=125).value
    dep_4_s_dob = worksheet.cell(row=i, column=126).value
    date_string = str(dep_4_s_dob)
    dep_4_dob = datetime_string.split(' ')[0]

    dep_5_firstname = worksheet.cell(row=i, column=141).value
    dep_5_middlename = worksheet.cell(row=i, column=142).value
    dep_5_lastname = worksheet.cell(row=i, column=143).value
    dep_5_ssn = worksheet.cell(row=i, column=144).value
    dep_4_s_dob = worksheet.cell(row=i, column=145).value
    date_string = str(dep_4_s_dob)
    dep_4_dob = datetime_string.split(' ')[0]

    dep_6_firstname = worksheet.cell(row=i, column=160).value
    dep_6_middlename = worksheet.cell(row=i, column=161).value
    dep_6_lastname = worksheet.cell(row=i, column=162).value
    dep_6_ssn = worksheet.cell(row=i, column=163).value
    dep_6_s_dob = worksheet.cell(row=i, column=164).value
    date_string = str(dep_6_s_dob)
    dep_6_dob = datetime_string.split(' ')[0]

    dep_7_firstname = worksheet.cell(row=i, column=179).value
    dep_7_middlename = worksheet.cell(row=i, column=180).value
    dep_7_lastname = worksheet.cell(row=i, column=181).value
    dep_7_ssn = worksheet.cell(row=i, column=182).value
    dep_7_s_dob = worksheet.cell(row=i, column=183).value
    date_string = str(dep_7_s_dob)
    dep_7_dob = datetime_string.split(' ')[0]

    dep_8_firstname = worksheet.cell(row=i, column=198).value
    dep_8_middlename = worksheet.cell(row=i, column=199).value
    dep_8_lastname = worksheet.cell(row=i, column=200).value
    dep_8_ssn = worksheet.cell(row=i, column=201).value
    dep_8_s_dob = worksheet.cell(row=i, column=202).value
    date_string = str(dep_8_s_dob)
    dep_8_dob = datetime_string.split(' ')[0]

    
    
    for j in dep:
        if j == None:
            pass
        else:
            dependent_count = dependent_count + 1

    print(write_row)

    employee_row = write_row+1
    if dependent_count > 0:
        total_increase = total_increase + dependent_count

    write_row = write_row + dependent_count
    write_row = write_row + 1


    wb2.save("D:\\PYTHON\\fileconversion\\David Horowitz\\Social_Circle_Nursing_and _Rehab\\ACAwise_FullService_Core_1094and1095-C_2024_V1.0.xlsx")        

        
        